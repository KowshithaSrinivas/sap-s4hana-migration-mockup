"""
Builds the Data Mapping Specification workbook: legacy field -> S/4HANA
Migration Cockpit target field, with transformation rules documented.
This is the artifact referenced in the JD as 'data mapping specifications'.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=13)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def write_sheet(name, title, columns, rows):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

    header_row = 3
    for c, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = WRAP

    for r, row in enumerate(rows, start=header_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = WRAP

    widths = [18, 16, 30, 16, 40, 22]
    for i, w in enumerate(widths[:len(columns)], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# Sheet 1: Material Master mapping
# ---------------------------------------------------------------------------
material_cols = ["Legacy Field", "Legacy Format/Example", "S/4HANA Target Field",
                  "Migration Object", "Transformation Rule", "Owner / Notes"]
material_rows = [
    ["Item Number", "ITEM-1234 (free text)", "MATNR", "Material", "Reformat to MAT-#### convention; validate uniqueness", "Data Migration Team"],
    ["Item Description", "Free text, mixed case", "MAKTX", "Material", "Trim whitespace; truncate to 40-char S/4 limit", "Data Migration Team"],
    ["Item Type", "Legacy codes: FG, SFG, RM", "MTART", "Material", "FG->FERT, SFG->HALB, RM->ROH (value mapping table)", "Business Process Owner"],
    ["Unit", "EA / ea / Each / PC (inconsistent)", "MEINS", "Material", "Standardize via UoM mapping table -> ISO-aligned code (PC/KG/L)", "Data Migration Team"],
    ["Plant Code", "1000, DE01, inconsistent whitespace", "WERKS", "Material", "Strip whitespace; map legacy plant codes to target plant via plant mapping table", "Global Process Owner"],
    ["Planning Method", "PD/VB/ND or blank", "DISMM", "Material (MRP view)", "Blank = hard error, must be resolved with business before load", "Business Process Owner"],
    ["Reorder Point", "Numeric or blank", "MINBE", "Material (MRP view)", "Required if DISMM=PD; flag for business input if missing", "Business Process Owner"],
    ["Safety Stock", "Numeric or blank", "EISBE", "Material (MRP view)", "Required if DISMM=PD; flag for business input if missing", "Business Process Owner"],
]

# ---------------------------------------------------------------------------
# Sheet 2: BOM mapping
# ---------------------------------------------------------------------------
bom_cols = ["Legacy Field", "Legacy Format/Example", "S/4HANA Target Field",
            "Migration Object", "Transformation Rule", "Owner / Notes"]
bom_rows = [
    ["BOM ID", "BOM-#### (legacy sequence)", "STLNR", "Bill of Material", "Regenerate as sequential S/4 BOM number; retain legacy ID as cross-reference", "Data Migration Team"],
    ["Parent Item", "Legacy item number", "Header Material (MATNR)", "Bill of Material", "Must resolve to a valid, already-migrated Material record", "Data Migration Team"],
    ["Component Item", "Legacy item number", "Component Material (MATNR)", "Bill of Material", "Referential integrity check against cleaned Material Master; exclude + flag if unresolved", "Data Migration Team"],
    ["Component Qty", "Numeric, 2 decimal", "MENGE", "Bill of Material", "No transformation; validate > 0", "Data Migration Team"],
    ["Component UoM", "EA/ea/KG (inconsistent)", "MEINS", "Bill of Material", "Standardize via shared UoM mapping table", "Data Migration Team"],
]

# ---------------------------------------------------------------------------
# Sheet 3: Work Center & Routing mapping
# ---------------------------------------------------------------------------
wcr_cols = ["Legacy Field", "Legacy Format/Example", "S/4HANA Target Field",
            "Migration Object", "Transformation Rule", "Owner / Notes"]
wcr_rows = [
    ["Work Center ID", "WC-### (legacy)", "ARBPL", "Work Center", "Retain legacy ID as cross-reference; validate uniqueness", "Data Migration Team"],
    ["Cost Center", "Mixed case, sometimes blank", "KOSTL", "Work Center", "Standardize to uppercase; blank = hard error (required for CO integration)", "Finance / Controlling"],
    ["Daily Capacity (hrs)", "Numeric or blank", "Capacity data", "Work Center", "Blank = warning; must confirm with plant before go-live", "Business Process Owner"],
    ["Routing ID", "RTG-#### (legacy)", "PLNNR", "Routing", "Regenerate as sequential S/4 routing number", "Data Migration Team"],
    ["Operation Sequence", "Numeric, 10-increment", "VORNR", "Routing", "No transformation; validate ascending, no gaps > business threshold", "Data Migration Team"],
    ["Operation Work Center", "Legacy WC ID", "ARBPL", "Routing", "Referential integrity check against cleaned Work Center list; exclude + flag if unresolved", "Data Migration Team"],
    ["Standard Time", "Minutes, 1 decimal", "Standard value (VGW00)", "Routing", "No transformation; validate > 0", "Data Migration Team"],
]

write_sheet("Material Master", "Data Mapping Specification — Material Master", material_cols, material_rows)
write_sheet("BOM", "Data Mapping Specification — Bill of Material", bom_cols, bom_rows)
write_sheet("Work Center & Routing", "Data Mapping Specification — Work Center & Routing", wcr_cols, wcr_rows)

# ---------------------------------------------------------------------------
# Cover sheet
# ---------------------------------------------------------------------------
cover = wb.create_sheet("Cover", 0)
cover["B2"] = "SAP S/4HANA Master Data Migration"
cover["B2"].font = Font(name="Arial", bold=True, size=16)
cover["B3"] = "Data Mapping Specification — Material & Production Master Data"
cover["B3"].font = Font(name="Arial", size=12, italic=True)
cover["B5"] = "Scope:"
cover["B5"].font = Font(name="Arial", bold=True, size=10)
cover["C5"] = "Material Master, Bill of Material, Work Center, Routing"
cover["C5"].font = BODY_FONT
cover["B6"] = "Prepared by:"
cover["B6"].font = Font(name="Arial", bold=True, size=10)
cover["C6"] = "Kowshitha Srinivas — self-directed migration readiness project"
cover["C6"].font = BODY_FONT
cover["B7"] = "Purpose:"
cover["B7"].font = Font(name="Arial", bold=True, size=10)
cover["C7"] = "Demonstrate mapping methodology and transformation rule design for SAP S/4HANA migration."
cover["C7"].font = BODY_FONT
cover.column_dimensions["B"].width = 14
cover.column_dimensions["C"].width = 60

out_path = os.path.join(os.path.dirname(__file__), "..", "mapping", "Data_Mapping_Specification.xlsx")
wb.save(out_path)
print(f"Saved: {os.path.abspath(out_path)}")
