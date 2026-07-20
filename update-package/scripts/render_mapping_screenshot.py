"""
Renders the Material Master mapping sheet as a clean image, reading the
actual content and colors from the real workbook (not re-typed data).
"""
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import os

BASE = os.path.dirname(__file__)
WB_PATH = os.path.join(BASE, "..", "mapping", "Data_Mapping_Specification.xlsx")
OUT_DIR = os.path.join(BASE, "..", "docs", "screenshots")

wb = openpyxl.load_workbook(WB_PATH)
ws = wb["Material Master"]

# Read header (row 3) and data rows (4 onward) directly from the real file
header = [c.value for c in ws[3]]
data_rows = []
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    if row[0] is None:
        continue
    data_rows.append(row)

n_cols = len(header)
n_rows = len(data_rows)

fig, ax = plt.subplots(figsize=(16, 0.5 * n_rows + 1.2))
ax.axis("off")

table = ax.table(
    cellText=data_rows,
    colLabels=header,
    loc="center",
    cellLoc="left",
    colLoc="left",
)
table.auto_set_font_size(False)
table.set_fontsize(10)
table.scale(1, 2.4)

col_widths = [0.11, 0.16, 0.15, 0.13, 0.32, 0.13]
for i, w in enumerate(col_widths):
    for r in range(n_rows + 1):
        table[(r, i)].set_width(w)

HEADER_BLUE = "#1F4E78"
for c in range(n_cols):
    cell = table[(0, c)]
    cell.set_facecolor(HEADER_BLUE)
    cell.set_text_props(color="white", fontweight="bold")

for r in range(1, n_rows + 1):
    for c in range(n_cols):
        table[(r, c)].set_facecolor("#F7F9FB" if r % 2 == 0 else "white")
        table[(r, c)].PAD = 0.02

ax.set_title("Data Mapping Specification — Material Master (excerpt)",
              fontsize=14, fontweight="bold", loc="left", pad=15)

plt.tight_layout()
plt.savefig(os.path.join(OUT_DIR, "mapping_spec_material_master.png"),
            dpi=150, bbox_inches="tight", facecolor="white")
plt.close()
print("Saved mapping_spec_material_master.png")
